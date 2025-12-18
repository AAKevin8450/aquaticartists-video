"""
Excel export utility for analysis results.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


def export_to_excel(job_data):
    """
    Convert analysis results to Excel format.

    Args:
        job_data: Dictionary containing job information and results

    Returns:
        BytesIO: Excel file as bytes
    """
    wb = Workbook()

    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_summary_sheet(ws_summary, job_data)

    # Create Data sheet based on analysis type
    ws_data = wb.create_sheet("Data")
    _create_data_sheet(ws_data, job_data)

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


def _create_summary_sheet(ws, job_data):
    """Create summary information sheet."""
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Title
    ws['A1'] = "Analysis Summary"
    ws['A1'].font = Font(size=16, bold=True)

    # Job Information
    row = 3
    info_fields = [
        ("Job ID:", job_data.get('job_id', 'N/A')),
        ("File Name:", job_data.get('file_name', 'N/A')),
        ("Analysis Type:", job_data.get('analysis_type_display', job_data.get('analysis_type', 'N/A'))),
        ("Status:", job_data.get('status', 'N/A')),
        ("Started At:", job_data.get('started_at', 'N/A')),
        ("Completed At:", job_data.get('completed_at', 'N/A')),
    ]

    for label, value in info_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    # Results count
    results = job_data.get('results', [])
    if isinstance(results, list):
        ws[f'A{row}'] = "Total Results:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = len(results)

    # Auto-size columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50


def _create_data_sheet(ws, job_data):
    """Create data sheet with analysis results."""
    analysis_type = job_data.get('analysis_type', '')
    results = job_data.get('results', [])

    if not results:
        ws['A1'] = "No results available"
        return

    # Determine format based on analysis type
    if 'label' in analysis_type.lower():
        _format_label_detection(ws, results)
    elif 'face' in analysis_type.lower() and 'search' not in analysis_type.lower():
        _format_face_detection(ws, results)
    elif 'celebrity' in analysis_type.lower():
        _format_celebrity_detection(ws, results)
    elif 'text' in analysis_type.lower():
        _format_text_detection(ws, results)
    elif 'moderation' in analysis_type.lower():
        _format_content_moderation(ws, results)
    elif 'person' in analysis_type.lower():
        _format_person_tracking(ws, results)
    elif 'segment' in analysis_type.lower():
        _format_segmentation(ws, results)
    else:
        _format_generic_results(ws, results)


def _format_label_detection(ws, results):
    """Format label detection results."""
    # Header
    headers = ["Timestamp (s)", "Label", "Confidence (%)", "Categories", "Instances Count"]
    _write_header_row(ws, headers)

    row = 2
    for item in results:
        timestamp = item.get('Timestamp', 0) / 1000  # Convert ms to seconds
        label = item.get('Label', {})

        categories = ', '.join([cat.get('Name', '') for cat in label.get('Categories', [])])
        instances_count = len(label.get('Instances', []))

        ws[f'A{row}'] = round(timestamp, 2)
        ws[f'B{row}'] = label.get('Name', '')
        ws[f'C{row}'] = round(label.get('Confidence', 0), 2)
        ws[f'D{row}'] = categories
        ws[f'E{row}'] = instances_count
        row += 1

    _auto_size_columns(ws, headers)


def _format_face_detection(ws, results):
    """Format face detection results."""
    headers = ["Timestamp (s)", "Confidence (%)", "Age Range", "Gender", "Emotions", "Smile", "Eyeglasses"]
    _write_header_row(ws, headers)

    row = 2
    for item in results:
        timestamp = item.get('Timestamp', 0) / 1000
        face = item.get('Face', {})

        age_range = face.get('AgeRange', {})
        age_str = f"{age_range.get('Low', 'N/A')}-{age_range.get('High', 'N/A')}"

        gender = face.get('Gender', {})
        gender_str = f"{gender.get('Value', 'N/A')} ({round(gender.get('Confidence', 0), 1)}%)"

        emotions = face.get('Emotions', [])
        emotion_str = ', '.join([f"{e.get('Type', '')} ({round(e.get('Confidence', 0), 1)}%)"
                                for e in emotions[:2]])  # Top 2 emotions

        smile = face.get('Smile', {})
        smile_str = f"{smile.get('Value', 'N/A')} ({round(smile.get('Confidence', 0), 1)}%)"

        eyeglasses = face.get('Eyeglasses', {})
        glasses_str = f"{eyeglasses.get('Value', 'N/A')} ({round(eyeglasses.get('Confidence', 0), 1)}%)"

        ws[f'A{row}'] = round(timestamp, 2)
        ws[f'B{row}'] = round(face.get('Confidence', 0), 2)
        ws[f'C{row}'] = age_str
        ws[f'D{row}'] = gender_str
        ws[f'E{row}'] = emotion_str
        ws[f'F{row}'] = smile_str
        ws[f'G{row}'] = glasses_str
        row += 1

    _auto_size_columns(ws, headers)


def _format_celebrity_detection(ws, results):
    """Format celebrity detection results."""
    headers = ["Timestamp (s)", "Celebrity Name", "Confidence (%)", "Match Confidence (%)", "URLs"]
    _write_header_row(ws, headers)

    row = 2
    for item in results:
        timestamp = item.get('Timestamp', 0) / 1000
        celebrity = item.get('Celebrity', {})

        name = celebrity.get('Name', 'Unknown')
        confidence = celebrity.get('Confidence', 0)
        match_confidence = celebrity.get('MatchConfidence', 0)

        urls = celebrity.get('Urls', [])
        urls_str = ', '.join(urls[:2]) if urls else 'N/A'

        ws[f'A{row}'] = round(timestamp, 2)
        ws[f'B{row}'] = name
        ws[f'C{row}'] = round(confidence, 2)
        ws[f'D{row}'] = round(match_confidence, 2)
        ws[f'E{row}'] = urls_str
        row += 1

    _auto_size_columns(ws, headers)


def _format_text_detection(ws, results):
    """Format text detection results."""
    headers = ["Timestamp (s)", "Detected Text", "Confidence (%)", "Type"]
    _write_header_row(ws, headers)

    row = 2
    for item in results:
        timestamp = item.get('Timestamp', 0) / 1000
        text_detection = item.get('TextDetection', {})

        ws[f'A{row}'] = round(timestamp, 2)
        ws[f'B{row}'] = text_detection.get('DetectedText', '')
        ws[f'C{row}'] = round(text_detection.get('Confidence', 0), 2)
        ws[f'D{row}'] = text_detection.get('Type', '')
        row += 1

    _auto_size_columns(ws, headers)


def _format_content_moderation(ws, results):
    """Format content moderation results."""
    headers = ["Timestamp (s)", "Label", "Confidence (%)", "Parent Category"]
    _write_header_row(ws, headers)

    row = 2
    for item in results:
        timestamp = item.get('Timestamp', 0) / 1000
        label = item.get('ModerationLabel', {})

        parent = label.get('ParentName', 'N/A')

        ws[f'A{row}'] = round(timestamp, 2)
        ws[f'B{row}'] = label.get('Name', '')
        ws[f'C{row}'] = round(label.get('Confidence', 0), 2)
        ws[f'D{row}'] = parent
        row += 1

    _auto_size_columns(ws, headers)


def _format_person_tracking(ws, results):
    """Format person tracking results."""
    headers = ["Timestamp (s)", "Person Index", "Confidence (%)"]
    _write_header_row(ws, headers)

    row = 2
    for item in results:
        timestamp = item.get('Timestamp', 0) / 1000
        person = item.get('Person', {})

        ws[f'A{row}'] = round(timestamp, 2)
        ws[f'B{row}'] = person.get('Index', 'N/A')
        ws[f'C{row}'] = round(person.get('Confidence', 0), 2)
        row += 1

    _auto_size_columns(ws, headers)


def _format_segmentation(ws, results):
    """Format shot/segment detection results."""
    headers = ["Type", "Timestamp (s)", "Duration (s)", "Confidence (%)"]
    _write_header_row(ws, headers)

    row = 2
    for item in results:
        segment_type = item.get('Type', 'N/A')
        timestamp = item.get('StartTimestampMillis', 0) / 1000
        duration = item.get('DurationMillis', 0) / 1000

        # Technical cue or shot detection
        confidence = 0
        if 'TechnicalCueSegment' in item:
            confidence = item['TechnicalCueSegment'].get('Confidence', 0)
        elif 'ShotSegment' in item:
            confidence = item['ShotSegment'].get('Confidence', 0)

        ws[f'A{row}'] = segment_type
        ws[f'B{row}'] = round(timestamp, 2)
        ws[f'C{row}'] = round(duration, 2)
        ws[f'D{row}'] = round(confidence, 2)
        row += 1

    _auto_size_columns(ws, headers)


def _format_generic_results(ws, results):
    """Generic format for unknown result types."""
    ws['A1'] = "Timestamp (s)"
    ws['B1'] = "Data"
    _write_header_row(ws, ["Timestamp (s)", "Data"])

    row = 2
    for item in results:
        timestamp = item.get('Timestamp', 0) / 1000
        ws[f'A{row}'] = round(timestamp, 2)
        ws[f'B{row}'] = str(item)
        row += 1

    _auto_size_columns(ws, ["Timestamp (s)", "Data"])


def _write_header_row(ws, headers):
    """Write and style header row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


def _auto_size_columns(ws, headers):
    """Auto-size columns based on content."""
    for col, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col)
        # Set minimum width based on header
        ws.column_dimensions[column_letter].width = max(len(header) + 2, 15)
